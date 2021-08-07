from openpyxl.styles.table import TableStyleElement
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook
import os



class searcher:

 
   def __init__(self):
       self.pesquisa_online()
       
       
   
   def pesquisa_online (self):

      self.navegador = webdriver.Chrome()
      self.navegador.minimize_window()
      time.sleep(3)
      os.system('cls') or None

      
      self.lista_nome =[]
      self.lista_preco =[]
      self.lista_percent =[]

      #Bitcoin
      self.navegador.get("https://br.tradingview.com/symbols/BTCBRL/")
      time.sleep(1)
      self.precoB = self.navegador.find_elements_by_xpath('/html/body/div[2]/div[6]/div/header/div/div[3]/div[1]/div/div/div/div[1]/div[1]')[0].text
      self.percentB = self.navegador.find_elements_by_xpath('/html/body/div[2]/div[6]/div/header/div/div[3]/div[1]/div/div/div/div[1]/div[3]/span[2]')[0].text
      print("===============  Bitcoin ===============\n" + "R$ " + self.precoB +"\n" + self.percentB)
      self.lista_nome.append('BitCoin')
      self.lista_preco.append(str(self.precoB))
      self.lista_percent.append(str(self.percentB))
      time.sleep(1)

      #Etherium
      self.navegador.get("https://br.tradingview.com/symbols/ETHBRL/")
      time.sleep(1)
      self.precoE = self.navegador.find_elements_by_xpath('/html/body/div[2]/div[6]/div/header/div/div[3]/div[1]/div/div/div/div[1]/div[1]/span')[0].text
      self.percentE = self.navegador.find_elements_by_xpath('/html/body/div[2]/div[6]/div/header/div/div[3]/div[1]/div/div/div/div[1]/div[3]/span[2]')[0].text
      print("===============  Etherium ===============\n" + "R$ " + self.precoE +"\n"+self.percentE )
      self.lista_nome.append('Etherium')
      self.lista_preco.append(str(self.precoE))
      self.lista_percent.append(str(self.percentE))
      time.sleep(1)

      print(f'Busca por cotação concluida com sucesso! \n{self.lista_nome[0]} : {self.lista_preco[0]} \n{self.lista_nome[1]} : {self.lista_preco[1]}')

      self.navegador.close()
      self.msg='pesquisa concluida'
     
 
   def criar_planilha (self):
      wb=Workbook()
      planilha = wb.worksheets[0]
      planilha.title = 'Análise Criptomoedas'
      planilha['A1'] = 'Nome'
      planilha['B1'] = 'Preço'
      index=2

      for nome , preco in zip(self.lista_nome,self.lista_preco):
         planilha.cell(column=1, row=index, value=nome)
         planilha.cell(column=2, row=index, value=preco)
         index+=1

      wb.save("Lista de precos.xlsx")
      print ('Salvo com sucesso')
   
